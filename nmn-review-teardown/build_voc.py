import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# (quote, brand, source, provenance, emotion, jtbd, objben, trigger, T,P,F,E,L, angle)
# T=time/place  P=physical sensation  F=former-self  E=embarrassing admission  L=plain unbranded
Y = "Y"; n = ""
R = []
a = R.append

# ---------------- ProHealth — Mark B, 4 stars, Jul 2026 (verbatim) ----------------
S = "ProHealth NMN Pro (Uthever)"; A = "Mark B · 4★ · Jul 2026 · Vine"; V = "Verbatim"
a(("After researching the ingredients and company behind this supplement I feel confident taking it.", S, A, V, "Pride", "Be the person who did the research", "BENEFIT · Credentialed brand", "Pre-purchase research", n,n,n,n,Y, "A7 Honest science"))
a(("I also looked at the broader science to see if NMN is just hype.", S, A, V, "Fear", "Verify it's real and safe", "OBJECTION · Category is hype", "Pre-purchase research", n,n,n,Y,Y, "A7 Honest science"))
a(("The honest answer is that the science is real but still emerging.", S, A, V, "Hope", "Verify it's real and safe", "OBJECTION · Evidence immature", "Pre-purchase research", n,n,n,n,Y, "A7 Honest science"))
a(("It is not snake oil but the anti aging claims are ahead of the evidence.", S, A, V, "Fear", "Verify it's real and safe", "OBJECTION · Marketing overreach", "Comparing brands", n,n,n,n,Y, "A7 Honest science"))
a(("A 2025 meta analysis found no clear benefit for muscle strength or metabolic markers.", S, A, V, "Fear", "Verify it's real and safe", "OBJECTION · Evidence immature", "Pre-purchase research", n,n,n,n,n, "A7 Honest science"))
a(("The FDA reversed its 2022 ban and now allows NMN as a supplement.", S, A, V, "Relief", "Verify it's real and safe", "OBJECTION · Was NMN banned?", "Pre-purchase research", Y,n,n,n,n, "A7 Honest science"))
a(("Perhaps in another year or two the effectiveness will be demonstrated in clinical testing.", S, A, V, "Hope", "Know whether it's working", "OBJECTION · Evidence immature", "Reorder decision", Y,n,n,n,Y, "A7 Honest science"))
a(("The marketing leans on benefits that are still being studied.", S, A, V, "Anger", "Not get ripped off", "OBJECTION · Marketing overreach", "Warning others", n,n,n,n,Y, "A7 Honest science"))
a(("TMG supports methylation which is important because NMN supplementation can deplete methyl groups.", S, A, V, "Pride", "Be the person who did the research", "BENEFIT · Cofactors included", "Reading the label", n,n,n,n,n, "A6 Retire the shelf"))
a(("Having all three in one bottle is convenient and actually costs less than buying them separately.", S, A, V, "Relief", "Simplify my routine", "BENEFIT · Consolidation", "Doing the price math", n,n,n,n,Y, "A6 Retire the shelf"))
a(("This is a well thought out stack.", S, A, V, "Pride", "Simplify my routine", "BENEFIT · Formula design", "Reading the label", n,n,n,n,Y, "A6 Retire the shelf"))
a(("The capsules are easy to swallow and have no unpleasant taste.", S, A, V, "Relief", "Protect a sensitive body", "BENEFIT · Sensory ease", "First dose", n,Y,n,n,Y, "A5 Gentle start"))
a(("I have noticed slight improved energy levels since starting.", S, A, V, "Hope", "Get my energy back", "BENEFIT · Mild energy", "Weeks 2–4", n,Y,n,n,Y, "A1 You won't feel it"))
a(("The 3 pack provides a 6 month supply at about 232 dollars which is premium pricing.", S, A, V, "Anger", "Justify the ongoing spend", "OBJECTION · Cost per day", "Doing the price math", n,n,n,n,Y, "A3 Do the math"))
a(("ProHealth has been around since 1988 and is a BBB A plus rated company.", S, A, V, "Relief", "Verify it's real and safe", "BENEFIT · Track record", "Comparing brands", Y,n,n,n,n, "A2 The receipt"))

# ---------------- ProHealth — kLr, 4 stars, Jun 2026 (verbatim) ----------------
A = "kLr · 4★ · Jun 2026 · Vine"
a(("I take a lot of supplements and am very focused on metabolically healthy aging as I am approaching 60.", S, A, V, "Fear", "Stay functional as I age", "BENEFIT · Identity fit", "Pre-purchase research", Y,n,Y,n,n, "A6 Retire the shelf"))
a(("They can add up, so it's important to me that they are not just effective, but a good value proposition and safe.", S, A, V, "Anger", "Justify the ongoing spend", "OBJECTION · Stack tax", "Doing the price math", n,n,n,n,Y, "A3 Do the math"))
a(("It's often hard to quantify changes in the short term on an individual basis.", S, A, V, "Hope", "Know whether it's working", "OBJECTION · No feedback loop", "Weeks 2–4", n,n,n,Y,Y, "A1 You won't feel it"))
a(("So many factors come into play in how I may feel on a particular day.", S, A, V, "Hope", "Know whether it's working", "OBJECTION · No feedback loop", "Weeks 2–4", n,Y,n,Y,Y, "A1 You won't feel it"))
a(("Have I exercised, did I sleep well, am I stressed, did I drink alcohol the day before?", S, A, V, "Shame", "Know whether it's working", "OBJECTION · No feedback loop", "Weeks 2–4", n,Y,n,Y,Y, "A1 You won't feel it"))
a(("I consider this a long-term health optimizer.", S, A, V, "Hope", "Stay functional as I age", "BENEFIT · Insurance frame", "Reorder decision", n,n,n,n,Y, "A1 You won't feel it"))
a(("I am very interested in giving my body the resources it needs to have the best outcome it can have as I age.", S, A, V, "Hope", "Stay functional as I age", "BENEFIT · Insurance frame", "Reorder decision", n,Y,n,n,Y, "A1 You won't feel it"))
a(("I am not new to these.", S, A, V, "Pride", "Be the person who did the research", "BENEFIT · Self-image", "Comparing brands", n,n,Y,n,Y, "A4 Dose is the product"))
a(("One might alternate both forms to get some of these benefits without totally breaking the budget.", S, A, V, "Shame", "Justify the ongoing spend", "OBJECTION · Cost per day", "Doing the price math", n,n,n,Y,Y, "A3 Do the math"))
a(("Lot tested, with public batch-specific certificates.", S, A, V, "Relief", "Verify it's real and safe", "BENEFIT · Batch COA", "Pre-purchase research", n,n,n,n,n, "A2 The receipt"))
a(("Lot number and Best By date is printed on the bottle.", S, A, V, "Relief", "Verify it's real and safe", "BENEFIT · Expiry printed", "Unboxing / first inspection", n,n,n,n,Y, "A2 The receipt"))
a(("It is confirmed made at a cGMP third-party certified facility.", S, A, V, "Relief", "Verify it's real and safe", "BENEFIT · Manufacturing", "Pre-purchase research", n,n,n,n,n, "A2 The receipt"))
a(("300 mg a day for 60 days in 66 adults, a 38% NAD+ increase by day 60.", S, A, V, "Pride", "Be the person who did the research", "BENEFIT · Human RCT", "Reading the label", Y,n,n,n,n, "A4 Dose is the product"))
a(("Results vary by interpretation, dose and duration.", S, A, V, "Fear", "Know whether it's working", "OBJECTION · Evidence immature", "Pre-purchase research", n,n,n,n,Y, "A7 Honest science"))
a(("Liposomal delivery is supposed to enhance absorption.", S, A, V, "Hope", "Verify it's real and safe", "OBJECTION · Absorption claim", "Reading the label", n,n,n,n,n, "A4 Dose is the product"))
a(("Depending on how much you wish to take, one bottle could last weeks or months.", S, A, V, "Fear", "Justify the ongoing spend", "OBJECTION · True supply length", "Doing the price math", n,n,n,n,Y, "A3 Do the math"))

# ---------------- ProHealth — Nancy Taylor, 3 stars, Jul 2026 ----------------
A = "Nancy Taylor · 3★ · Jul 2026"
a(("I kind of had high hopes for this longevity supplement.", S, A, V, "Grief", "Know whether it's working", "OBJECTION · Unmet expectation", "Weeks 2–4", n,n,n,Y,Y, "A1 You won't feel it"))
a(("I'm not feeling any different energy wise.", S, A, V, "Grief", "Get my energy back", "OBJECTION · Felt nothing", "Weeks 2–4", n,Y,n,n,Y, "A1 You won't feel it"))
a(("But maybe I need to give it more time?", S, A, V, "Hope", "Know whether it's working", "OBJECTION · Timeline unclear", "Weeks 2–4", n,n,n,Y,Y, "A1 You won't feel it"))
a(("It's not inexpensive so I probably won't be ordering again.", S, A, V, "Anger", "Justify the ongoing spend", "OBJECTION · Cost per day", "Reorder decision", n,n,n,n,Y, "A3 Do the math"))
a(("Was hoping to notice a change while taking.", S, A, V, "Grief", "Know whether it's working", "OBJECTION · Felt nothing", "Weeks 2–4", n,Y,n,Y,Y, "A1 You won't feel it"))

# ---------------- ProHealth — Matthew B, 1 star, Mar 2022 ----------------
A = "Matthew B · 1★ · Mar 2022"
a(("As recommended by the bottle I took two pills around 8am.", S, A, V, "Anger", "Protect a sensitive body", "OBJECTION · Dose guidance", "First dose", Y,n,n,n,Y, "A5 Gentle start"))
a(("I was unable to fall asleep until 6am the following morning.", S, A, V, "Anger", "Sleep through the night", "OBJECTION · Insomnia", "First dose", Y,Y,n,n,Y, "A5 Gentle start"))
a(("I experienced an increase in energy, an increase in heart rate, and anxiousness.", S, A, V, "Fear", "Protect a sensitive body", "OBJECTION · Over-stimulation", "First dose", n,Y,n,n,Y, "A5 Gentle start"))
a(("I do not drink caffeinated drinks as I am sensitive to caffeine.", S, A, V, "Fear", "Protect a sensitive body", "BENEFIT · Segment tell", "Pre-purchase research", n,Y,n,n,Y, "A5 Gentle start"))
a(("I suspect this is just a caffeine pill.", S, A, V, "Anger", "Not get ripped off", "OBJECTION · Is it real?", "Abandonment / refund", n,n,n,n,Y, "A2 The receipt"))
a(("Buyer beware.", S, A, V, "Anger", "Not get ripped off", "OBJECTION · Is it real?", "Warning others", n,n,n,n,Y, "A2 The receipt"))
a(("Will not be taking this again.", S, A, V, "Anger", "Not get ripped off", "OBJECTION · Abandonment", "Abandonment / refund", n,n,n,n,Y, "A5 Gentle start"))

# ---------------- ProHealth — Rick, 3 stars, Sep 2021 ----------------
A = "Rick · 3★ · Sep 2021"
a(("I have bought 12 bottles of this and I have not felt much difference.", S, A, V, "Grief", "Know whether it's working", "OBJECTION · Felt nothing", "Month 2+", n,Y,n,Y,Y, "A1 You won't feel it"))
a(("I don't know where this stuff is stored or how long it's been on the shelves.", S, A, V, "Fear", "Verify it's real and safe", "OBJECTION · Degradation", "Unboxing / first inspection", Y,n,n,n,Y, "A2 The receipt"))
a(("My results have been lackluster.", S, A, V, "Grief", "Know whether it's working", "OBJECTION · Felt nothing", "Month 2+", n,n,n,n,Y, "A1 You won't feel it"))
a(("I took Niagen brand of NR before trying NMN and I was noticing some cognitive benefit.", S, A, V, "Hope", "Think clearly again", "OBJECTION · NR worked better", "Comparing brands", n,Y,Y,n,n, "A4 Dose is the product"))
a(("I take lots of supplements so maybe there was no deficiency that needed to be addressed.", S, A, V, "Shame", "Know whether it's working", "OBJECTION · Stack tax", "Month 2+", n,n,n,Y,Y, "A6 Retire the shelf"))
a(("I bought some more expensive third party tested NMN and again I notice very little difference.", S, A, V, "Grief", "Know whether it's working", "OBJECTION · Felt nothing", "Reorder decision", n,n,Y,Y,Y, "A1 You won't feel it"))
a(("I didn't feel much.", S, A, V, "Grief", "Get my energy back", "OBJECTION · Felt nothing", "Weeks 2–4", n,Y,n,n,Y, "A1 You won't feel it"))

# ---------------- ProHealth — other verbatims ----------------
A = "Amazon Reviewer · 3★ · Nov 2022"
a(("After hundreds of hours research on NMN and NR, I decided to buy this product.", S, A, V, "Pride", "Be the person who did the research", "BENEFIT · Self-image", "Pre-purchase research", n,n,n,n,n, "A4 Dose is the product"))
a(("I do not feel energy boost or anything like that.", S, A, V, "Grief", "Get my energy back", "OBJECTION · Felt nothing", "Weeks 2–4", n,Y,n,n,Y, "A1 You won't feel it"))
a(("Calm but hard to focus and daydreaming.", S, A, V, "Fear", "Think clearly again", "OBJECTION · Paradox effect", "Week 1", n,Y,n,n,Y, "A5 Gentle start"))
a(("I have many vivid dreams with light headache when I wake up.", S, A, V, "Fear", "Sleep through the night", "OBJECTION · Side effect", "Week 1", Y,Y,n,n,Y, "A5 Gentle start"))
a(("Not sure this product contains real NMN or something else in it.", S, A, V, "Fear", "Not get ripped off", "OBJECTION · Is it real?", "Weeks 2–4", n,n,n,n,Y, "A2 The receipt"))
a(("Be careful.", S, A, V, "Fear", "Not get ripped off", "OBJECTION · Is it real?", "Warning others", n,n,n,n,Y, "A2 The receipt"))

A = "CHAN MENG CHIA · 4★ · Jul 2022"
a(("I've tried it for less than a month and I haven't felt anything special yet.", S, A, V, "Hope", "Know whether it's working", "OBJECTION · Felt nothing", "Weeks 2–4", Y,Y,n,n,Y, "A1 You won't feel it"))
a(("I have used similar products before, I felt obviously the change in my body.", S, A, V, "Grief", "Get my energy back", "OBJECTION · Competitor worked", "Comparing brands", n,Y,Y,n,Y, "A1 You won't feel it"))

A = "Ms12587 · 1★ · Oct 2021 · Singapore"
a(("In each bottle, I found several capsules that are broken.", S, A, V, "Anger", "Verify it's real and safe", "OBJECTION · Physical quality", "Unboxing / first inspection", n,Y,n,n,Y, "A2 The receipt"))
a(("The broken capsules released the powder and all the other capsules tasted salty.", S, A, V, "Anger", "Verify it's real and safe", "OBJECTION · Physical quality", "First dose", n,Y,n,n,Y, "A2 The receipt"))
a(("Some capsules are intact but empty or only half full.", S, A, V, "Anger", "Not get ripped off", "OBJECTION · Short-filled", "Unboxing / first inspection", n,Y,n,n,Y, "A2 The receipt"))
a(("This has made me seriously wonder the overall quality of the product.", S, A, V, "Fear", "Verify it's real and safe", "OBJECTION · Trust collapse", "Unboxing / first inspection", n,n,n,n,Y, "A2 The receipt"))
a(("They claim high-standard QC. Based on what I see, this is not what they have claimed.", S, A, V, "Anger", "Not get ripped off", "OBJECTION · Claim vs reality", "Warning others", n,n,n,n,Y, "A2 The receipt"))

A = "Tall Young · 1★ · Jul 2022"
a(("The taste is so awful, smells like chemicals.", S, A, V, "Anger", "Protect a sensitive body", "OBJECTION · Smell / taste", "First dose", n,Y,n,n,Y, "A5 Gentle start"))
a(("Gave me an allergic reaction.", S, A, V, "Fear", "Protect a sensitive body", "OBJECTION · Adverse reaction", "First dose", n,Y,n,n,Y, "A5 Gentle start"))
a(("Cannot be returned.", S, A, V, "Anger", "Not get ripped off", "OBJECTION · No refund", "Abandonment / refund", n,n,n,n,Y, "A3 Do the math"))
a(("Do not purchase this product.", S, A, V, "Anger", "Not get ripped off", "OBJECTION · Trust collapse", "Warning others", n,n,n,n,Y, "A2 The receipt"))

A = "Vine reviewer · 4★ · Jun 2026"
a(("I've become more interested in healthy aging and longevity as I get older.", S, A, V, "Fear", "Stay functional as I age", "BENEFIT · Identity fit", "Pre-purchase research", n,n,Y,n,Y, "A6 Retire the shelf"))
a(("My energy levels have been a little more consistent and I seem to have better stamina.", S, A, V, "Relief", "Get my energy back", "BENEFIT · Subtle energy", "Weeks 2–4", n,Y,Y,n,Y, "A1 You won't feel it"))
a(("The bottle contains as little as a 15-day supply.", S, A, V, "Anger", "Justify the ongoing spend", "OBJECTION · True supply length", "Doing the price math", n,n,n,n,Y, "A3 Do the math"))
a(("The bottle felt and looked practically empty when I first opened it.", S, A, V, "Anger", "Not get ripped off", "OBJECTION · Fill level", "Unboxing / first inspection", n,Y,n,n,Y, "A3 Do the math"))
a(("There were 30 capsules. I counted them to be sure.", S, A, V, "Fear", "Not get ripped off", "OBJECTION · Fill level", "Unboxing / first inspection", n,n,n,Y,Y, "A3 Do the math"))
a(("This product can get expensive if you plan to take it long term.", S, A, V, "Fear", "Justify the ongoing spend", "OBJECTION · Cost per day", "Reorder decision", n,n,n,n,Y, "A3 Do the math"))
a(("I haven't experienced any side effects, stomach upset, or other issues.", S, A, V, "Relief", "Protect a sensitive body", "BENEFIT · Tolerability", "Weeks 2–4", n,Y,n,n,Y, "A5 Gentle start"))
a(("Aging gracefully, which I am hoping to do.", S, A, V, "Hope", "Slow visible aging", "BENEFIT · Identity fit", "Pre-purchase research", n,n,n,Y,Y, "A6 Retire the shelf"))

A = "Kyle · 4★ · Jun 2026 · Vine"
a(("Be patient.", S, A, V, "Hope", "Know whether it's working", "BENEFIT · Realistic timeline", "Week 1", n,n,n,n,Y, "A1 You won't feel it"))
a(("It takes a little longer to take effect since your body needs to digest them first.", S, A, V, "Hope", "Know whether it's working", "OBJECTION · Timeline unclear", "Week 1", n,Y,n,n,Y, "A1 You won't feel it"))
a(("The capsules aren't too large nor do they have any unpleasant taste or scent.", S, A, V, "Relief", "Protect a sensitive body", "BENEFIT · Sensory ease", "First dose", n,Y,n,n,Y, "A5 Gentle start"))
a(("A really nice boost of energy, and it's sustained, so no crash later in the day.", S, A, V, "Relief", "Get my energy back", "BENEFIT · Clean energy", "Weeks 2–4", Y,Y,n,n,Y, "A1 You won't feel it"))

# ---------------- ProHealth powder — Jackie Waldrop ----------------
S2 = "ProHealth NMN Powder"; A = "Jackie Waldrop · 4★ · Jan 2023"
a(("I read where people kept saying they did not get a spoon but mine was in the bag.", S2, A, V, "Relief", "Not get ripped off", "BENEFIT · Fear resolved", "Unboxing / first inspection", n,n,n,n,Y, "A2 The receipt"))
a(("The taste is kind of twangy. Not bad.", S2, A, V, "Relief", "Protect a sensitive body", "BENEFIT · Taste acceptable", "First dose", n,Y,n,n,Y, "A5 Gentle start"))
a(("It disappears in the water with no cloud and is clear, easy to drink.", S2, A, V, "Relief", "Protect a sensitive body", "BENEFIT · Dissolves clean", "First dose", n,Y,n,n,Y, "A5 Gentle start"))
a(("My husband thought I gave him just a glass of water.", S2, A, V, "Relief", "Care for my spouse", "BENEFIT · Dissolves clean", "First dose", n,Y,n,n,Y, "A5 Gentle start"))
a(("He has COPD and I have IBS. Wanted to try it for those two things.", S2, A, V, "Hope", "Care for my spouse", "OBJECTION · Disease hope [DO NOT USE]", "Pre-purchase research", n,Y,n,n,Y, "§07 Guardrail — do not use"))
a(("I had a bad reaction the first day and had cramps in my stomach.", S2, A, V, "Fear", "Protect a sensitive body", "OBJECTION · Gut reaction", "First dose", Y,Y,n,n,Y, "A5 Gentle start"))
a(("I was afraid to try it again.", S2, A, V, "Fear", "Protect a sensitive body", "OBJECTION · Gut reaction", "Week 1", n,Y,n,Y,Y, "A5 Gentle start"))
a(("By the third day he said I think it is helping me, so I started back taking it.", S2, A, V, "Hope", "Care for my spouse", "BENEFIT · Push through week 1", "Week 1", Y,n,n,n,Y, "A5 Gentle start"))
a(("Keeping a log of our vitals.", S2, A, V, "Hope", "Know whether it's working", "BENEFIT · Wants a scoreboard", "Week 1", n,n,n,n,Y, "A1 You won't feel it"))
a(("Hoping that soon he won't have to take them any more and just the NMN.", S2, A, V, "Hope", "Care for my spouse", "OBJECTION · Drug replacement [DO NOT USE]", "Month 2+", n,n,n,n,Y, "§07 Guardrail — do not use"))

# ---------------- Nutricost NAD+ ----------------
S = "Nutricost NAD+ 1,000mg"; A = "Listing summary"; P = "Paraphrase — not quotable"
a(("NAD+ as a molecule is too large to be effectively absorbed through the digestive system.", S, A, P, "Fear", "Verify it's real and safe", "OBJECTION · Bioavailability", "Pre-purchase research", n,n,n,n,n, "A4 Dose is the product"))
a(("Precursors like NMN or NR may be more effective for raising NAD+ levels.", S, A, P, "Hope", "Be the person who did the research", "BENEFIT · NMN over NAD+", "Comparing brands", n,n,n,n,n, "A4 Dose is the product"))
a(("Too strong.", S, A, "Quoted fragment", "Fear", "Protect a sensitive body", "OBJECTION · Over-stimulation", "First dose", n,Y,n,n,Y, "A5 Gentle start"))
a(("Start with just 1 capsule to assess tolerance.", S, A, P, "Fear", "Protect a sensitive body", "BENEFIT · Self-invented protocol", "First dose", n,n,n,n,Y, "A5 Gentle start"))
a(("Does 1,000mg per serving exceed safe daily intake guidelines?", S, A, P, "Fear", "Verify it's real and safe", "OBJECTION · Dose too high", "Reading the label", n,n,n,n,Y, "A4 Dose is the product"))
a(("Should dosing be adjusted based on body weight?", S, A, P, "Fear", "Verify it's real and safe", "OBJECTION · Dose guidance", "Reading the label", n,n,n,n,Y, "A4 Dose is the product"))
a(("Stomach aches, cramping, and in some cases vomiting.", S, A, P, "Fear", "Protect a sensitive body", "OBJECTION · Gut reaction", "Week 1", n,Y,n,n,Y, "A5 Gentle start"))

# ---------------- Nutricost NMN 500 ----------------
S = "Nutricost NMN 500mg"; A = "Listing summary"
a(("Rancid.", S, A, "Quoted fragment", "Anger", "Protect a sensitive body", "OBJECTION · Smell", "Unboxing / first inspection", n,Y,n,n,Y, "A5 Gentle start"))
a(("Smells like interior paint.", S, A, "Quoted fragment", "Anger", "Protect a sensitive body", "OBJECTION · Smell", "Unboxing / first inspection", n,Y,n,n,Y, "A5 Gentle start"))
a(("Did not feel any changes.", S, A, "Quoted fragment", "Grief", "Know whether it's working", "OBJECTION · Felt nothing", "Weeks 2–4", n,Y,n,n,Y, "A1 You won't feel it"))
a(("Not sure if it's working.", S, A, "Quoted fragment", "Hope", "Know whether it's working", "OBJECTION · No feedback loop", "Weeks 2–4", n,n,n,Y,Y, "A1 You won't feel it"))
a(("Not for me.", S, A, "Quoted fragment", "Shame", "Protect a sensitive body", "OBJECTION · Poor fit", "Abandonment / refund", n,n,n,Y,Y, "A5 Gentle start"))
a(("Felt more worn out and less alert after a week of use.", S, A, P, "Grief", "Get my energy back", "OBJECTION · Paradox effect", "Week 1", Y,Y,Y,n,Y, "A5 Gentle start"))

# ---------------- Double Wood NAD+ ----------------
S = "Double Wood NAD+ 500mg"; A = "Listing summary"
a(("Crackhead energy.", S, A, "Quoted fragment", "Fear", "Protect a sensitive body", "OBJECTION · Over-stimulation", "First dose", n,Y,n,Y,Y, "A5 Gentle start"))
a(("Not sure if we were supposed to feel anything.", S, A, "Quoted fragment", "Shame", "Know whether it's working", "OBJECTION · No feedback loop", "Weeks 2–4", n,Y,n,Y,Y, "A1 You won't feel it"))
a(("Misleading, waste of money.", S, A, "Quoted fragment", "Anger", "Not get ripped off", "OBJECTION · Marketing overreach", "Abandonment / refund", n,n,n,n,Y, "A7 Honest science"))
a(("Tired, lethargic and drained. The opposite of the expected energy boost.", S, A, P, "Grief", "Get my energy back", "OBJECTION · Paradox effect", "Week 1", n,Y,Y,n,Y, "A5 Gentle start"))
a(("Irritability, anxiety and jitteriness initially.", S, A, P, "Fear", "Protect a sensitive body", "OBJECTION · Over-stimulation", "Week 1", n,Y,n,n,Y, "A5 Gentle start"))
a(("Breast tenderness as an unexpected side effect.", S, A, P, "Fear", "Protect a sensitive body", "OBJECTION · Unexpected effect", "Weeks 2–4", n,Y,n,Y,Y, "A5 Gentle start"))
a(("Even after 4 to 5 months of use, no changes in energy, mood or well-being.", S, A, P, "Grief", "Know whether it's working", "OBJECTION · Felt nothing", "Month 2+", Y,Y,n,n,Y, "A1 You won't feel it"))
a(("Is oral NAD+ even effectively absorbed by the body?", S, A, P, "Fear", "Verify it's real and safe", "OBJECTION · Bioavailability", "Pre-purchase research", n,n,n,n,Y, "A4 Dose is the product"))

# ---------------- 4Well NAD+ ----------------
S = "4Well NAD Plus 1400mg"; A = "Listing summary"
a(("Very acidic.", S, A, "Quoted fragment", "Fear", "Protect a sensitive body", "OBJECTION · Gut reaction", "First dose", n,Y,n,n,Y, "A5 Gentle start"))
a(("This supplement did nothing.", S, A, "Quoted fragment", "Anger", "Know whether it's working", "OBJECTION · Felt nothing", "Month 2+", n,n,n,n,Y, "A1 You won't feel it"))
a(("Utterly useless.", S, A, "Quoted fragment", "Anger", "Not get ripped off", "OBJECTION · Bioavailability", "Abandonment / refund", n,n,n,n,Y, "A4 Dose is the product"))
a(("It does not survive digestive acids.", S, A, P, "Fear", "Verify it's real and safe", "OBJECTION · Bioavailability", "Pre-purchase research", n,Y,n,n,Y, "A4 Dose is the product"))
a(("Zero results even after taking two full jars over several months.", S, A, P, "Grief", "Know whether it's working", "OBJECTION · Felt nothing", "Month 2+", n,n,n,n,Y, "A1 You won't feel it"))
a(("Nausea, vomiting, diarrhea and excessive gas.", S, A, P, "Fear", "Protect a sensitive body", "OBJECTION · Gut reaction", "Week 1", n,Y,n,n,Y, "A5 Gentle start"))
a(("The seller did not allow returns.", S, A, P, "Anger", "Not get ripped off", "OBJECTION · No refund", "Abandonment / refund", n,n,n,n,Y, "A3 Do the math"))

# ---------------- ForestLeaf ----------------
S = "ForestLeaf NAD+ 500mg"; A = "Listing summary"
a(("Did not see any results.", S, A, "Quoted fragment", "Grief", "Know whether it's working", "OBJECTION · Felt nothing", "Weeks 2–4", n,n,n,n,Y, "A1 You won't feel it"))
a(("Misleading.", S, A, "Quoted fragment", "Anger", "Not get ripped off", "OBJECTION · Marketing overreach", "Abandonment / refund", n,n,n,n,Y, "A7 Honest science"))
a(("A quality made product, but no difference after 4 weeks.", S, A, "Quoted fragment", "Grief", "Know whether it's working", "OBJECTION · Felt nothing", "Weeks 2–4", Y,n,n,n,Y, "A1 You won't feel it"))
a(("The claim of immediate absorption for faster results is misleading.", S, A, P, "Anger", "Not get ripped off", "OBJECTION · Marketing overreach", "Weeks 2–4", n,n,n,n,Y, "A7 Honest science"))
a(("Is oral NAD+ as effective as IV or sublingual?", S, A, P, "Fear", "Verify it's real and safe", "OBJECTION · Bioavailability", "Comparing brands", n,n,n,n,n, "A4 Dose is the product"))

# ---------------- Tru Niagen ----------------
S = "Tru Niagen NR 300mg"; A = "Listing summary / Q&A"
a(("Too expensive.", S, A, "Quoted fragment", "Anger", "Justify the ongoing spend", "OBJECTION · Cost per day", "Doing the price math", n,n,n,n,Y, "A3 Do the math"))
a(("Did nothing for me.", S, A, "Quoted fragment", "Grief", "Know whether it's working", "OBJECTION · Felt nothing", "Month 2+", n,n,n,n,Y, "A1 You won't feel it"))
a(("I possibly felt worse.", S, A, "Quoted fragment", "Grief", "Get my energy back", "OBJECTION · Paradox effect", "Weeks 2–4", n,Y,Y,Y,Y, "A5 Gentle start"))
a(("Waste of money.", S, A, "Quoted fragment", "Anger", "Not get ripped off", "OBJECTION · Cost per day", "Abandonment / refund", n,n,n,n,Y, "A3 Do the math"))
a(("Do the capsules actually contain the stated 300mg?", S, A, P, "Fear", "Not get ripped off", "OBJECTION · Is it real?", "Reading the label", n,n,n,n,Y, "A2 The receipt"))
a(("Concerns about shelf life and degradation if not refrigerated.", S, A, P, "Fear", "Verify it's real and safe", "OBJECTION · Degradation", "Unboxing / first inspection", n,n,n,n,n, "A2 The receipt"))
a(("Nausea, stomach pain, dizziness, lightheadedness and brain fog.", S, A, P, "Fear", "Protect a sensitive body", "OBJECTION · Side effect cluster", "Week 1", n,Y,n,n,Y, "A5 Gentle start"))
a(("Insomnia and sleep disruption, the opposite of the expected benefit.", S, A, P, "Anger", "Sleep through the night", "OBJECTION · Paradox effect", "Week 1", n,Y,n,n,Y, "A5 Gentle start"))
a(("Palpitations and anxiety after taking it.", S, A, P, "Fear", "Protect a sensitive body", "OBJECTION · Over-stimulation", "First dose", n,Y,n,n,Y, "A5 Gentle start"))

# ---------------- Nutricost NADH ----------------
S = "Nutricost NADH 20mg"; A = "Listing summary"
a(("The jury is out on effectiveness.", S, A, "Quoted fragment", "Hope", "Know whether it's working", "OBJECTION · No feedback loop", "Weeks 2–4", n,n,n,n,Y, "A1 You won't feel it"))
a(("I am not sure we are noticing any difference.", S, A, "Quoted fragment", "Hope", "Know whether it's working", "OBJECTION · No feedback loop", "Weeks 2–4", n,n,n,Y,Y, "A1 You won't feel it"))
a(("Very sick for several days with diarrhea and stomach cramps.", S, A, P, "Fear", "Protect a sensitive body", "OBJECTION · Gut reaction", "Week 1", Y,Y,n,n,Y, "A5 Gentle start"))
a(("It significantly lowered blood pressure rather than boosting energy.", S, A, P, "Fear", "Protect a sensitive body", "OBJECTION · Unexpected effect", "Weeks 2–4", n,Y,n,n,Y, "§07 Guardrail — do not use"))
a(("It takes several weeks before any benefit is felt.", S, A, P, "Hope", "Know whether it's working", "OBJECTION · Timeline unclear", "Weeks 2–4", n,Y,n,n,Y, "A1 You won't feel it"))

# ---------------- Renue LIPO NMN ----------------
S = "Renue By Science LIPO NMN"; A = "Listing summary"
a(("Barely functional.", S, A, "Quoted fragment", "Grief", "Think clearly again", "OBJECTION · Paradox effect", "Weeks 2–4", n,Y,n,n,Y, "A5 Gentle start"))
a(("It significantly worsened my fatigue and brain fog.", S, A, P, "Anger", "Get my energy back", "OBJECTION · Paradox effect", "Weeks 2–4", n,Y,Y,n,Y, "A5 Gentle start"))
a(("Bloating, heaviness and decreased energy.", S, A, P, "Anger", "Get my energy back", "OBJECTION · Paradox effect", "Week 1", n,Y,n,n,Y, "A5 Gentle start"))
a(("At 65 I needed more energy, not less.", S, A, P, "Grief", "Stay functional as I age", "OBJECTION · Paradox effect", "Abandonment / refund", n,Y,Y,n,Y, "A5 Gentle start"))
a(("I assumed my body needed to adjust, but the symptoms persisted.", S, A, P, "Grief", "Protect a sensitive body", "OBJECTION · Side effect", "Weeks 2–4", n,Y,n,Y,Y, "A5 Gentle start"))

# ---------------- Renue NMN Powder ----------------
S = "Renue By Science NMN Powder"; A = "Listing summary"
a(("No way to take it with such an awful bitter and disgusting taste.", S, A, "Quoted fragment", "Anger", "Protect a sensitive body", "OBJECTION · Taste", "First dose", n,Y,n,n,Y, "A5 Gentle start"))
a(("It doesn't really dissolve so much as become a sludge.", S, A, "Quoted fragment", "Anger", "Protect a sensitive body", "OBJECTION · Texture", "First dose", n,Y,n,n,Y, "A5 Gentle start"))
a(("Feel or look no different.", S, A, "Quoted fragment", "Grief", "Slow visible aging", "OBJECTION · Felt nothing", "Weeks 2–4", n,Y,Y,n,Y, "A1 You won't feel it"))
a(("I have not seen any results from this product yet.", S, A, "Quoted fragment", "Hope", "Know whether it's working", "OBJECTION · Felt nothing", "Weeks 2–4", n,n,n,n,Y, "A1 You won't feel it"))
a(("The taste was so bad I couldn't finish a single dose.", S, A, P, "Anger", "Protect a sensitive body", "OBJECTION · Taste", "First dose", n,Y,n,Y,Y, "A5 Gentle start"))
a(("Nausea began after switching to a new label and batch.", S, A, P, "Fear", "Verify it's real and safe", "OBJECTION · Batch consistency", "Reorder decision", n,Y,Y,n,Y, "A2 The receipt"))

# ---------------- Renue cross-product ----------------
S = "Renue By Science (multiple SKUs)"; A = "Listing summary"
a(("Only a manufacture date is provided, with no expiration date.", S, A, P, "Fear", "Verify it's real and safe", "OBJECTION · No expiry", "Unboxing / first inspection", n,n,n,n,Y, "A2 The receipt"))
a(("Spendy.", S, A, "Quoted fragment", "Anger", "Justify the ongoing spend", "OBJECTION · Cost per day", "Doing the price math", n,n,n,n,Y, "A3 Do the math"))
a(("Studies use 500 to 1000mg a day and standard servings fall short.", S, A, P, "Anger", "Be the person who did the research", "OBJECTION · Underdosed", "Reading the label", n,n,n,n,n, "A4 Dose is the product"))
a(("Results are subtle and hard to attribute directly to the supplement.", S, A, P, "Hope", "Know whether it's working", "OBJECTION · No feedback loop", "Weeks 2–4", n,Y,n,n,Y, "A1 You won't feel it"))

# ---------------- Renue SLC Enteric ----------------
S = "Renue By Science SLC Enteric NMN 250mg"; A = "Listing summary"
a(("Did nothing.", S, A, "Quoted fragment", "Anger", "Know whether it's working", "OBJECTION · Felt nothing", "Weeks 2–4", n,n,n,n,Y, "A1 You won't feel it"))
a(("Can see no benefit yet.", S, A, "Quoted fragment", "Hope", "Know whether it's working", "OBJECTION · Felt nothing", "Weeks 2–4", n,n,n,n,Y, "A1 You won't feel it"))
a(("The gradual, subtle effects make it hard to know if it's actually working.", S, A, P, "Hope", "Know whether it's working", "OBJECTION · No feedback loop", "Weeks 2–4", n,Y,n,n,Y, "A1 You won't feel it"))
a(("250mg per serving is not enough. I prefer 500 to 1000mg daily.", S, A, P, "Anger", "Be the person who did the research", "OBJECTION · Underdosed", "Reading the label", n,n,n,n,n, "A4 Dose is the product"))
a(("Is enteric-coated NMN actually better than regular NMN?", S, A, P, "Fear", "Not get ripped off", "OBJECTION · Absorption claim", "Comparing brands", n,n,n,n,n, "A4 Dose is the product"))
a(("I got it via Vine and wouldn't have purchased it at $25.95.", S, A, P, "Shame", "Justify the ongoing spend", "OBJECTION · Cost per day", "Reorder decision", n,n,n,Y,Y, "A3 Do the math"))

# ---------------- Renue Spermidine ----------------
S = "Renue By Science Liposomal Spermidine"; A = "Listing summary"
a(("At 56 cents a capsule it's far pricier than competitors at 33 to 37 cents.", S, A, P, "Anger", "Justify the ongoing spend", "OBJECTION · Cost per day", "Doing the price math", n,n,n,n,Y, "A3 Do the math"))
a(("Spermidine is naturally water-soluble and doesn't need liposomal encapsulation.", S, A, P, "Anger", "Not get ripped off", "OBJECTION · Absorption claim", "Comparing brands", n,n,n,n,n, "A4 Dose is the product"))
a(("Research shows 87 to 96% is already absorbed without liposomal delivery.", S, A, P, "Anger", "Be the person who did the research", "OBJECTION · Absorption claim", "Comparing brands", n,n,n,n,n, "A4 Dose is the product"))
a(("I'm not sure how much it is helping me.", S, A, "Quoted fragment", "Hope", "Know whether it's working", "OBJECTION · No feedback loop", "Weeks 2–4", n,n,n,Y,Y, "A1 You won't feel it"))
a(("Long air pockets under the label on the bottle.", S, A, P, "Fear", "Verify it's real and safe", "OBJECTION · Physical quality", "Unboxing / first inspection", n,Y,n,n,Y, "A2 The receipt"))
a(("The bottle arrived with a broken seal and open cap.", S, A, P, "Fear", "Verify it's real and safe", "OBJECTION · Physical quality", "Unboxing / first inspection", n,Y,n,n,Y, "A2 The receipt"))

# ---------------- Renue Chewables ----------------
S = "Renue By Science NMN Chewables 125mg"; A = "Listing summary"
a(("Chewing them is a no go for me.", S, A, "Quoted fragment", "Anger", "Protect a sensitive body", "OBJECTION · Taste", "First dose", n,Y,n,n,Y, "A5 Gentle start"))
a(("The lemon taste is offset by some other underlying taste.", S, A, "Quoted fragment", "Anger", "Protect a sensitive body", "OBJECTION · Taste", "First dose", n,Y,n,n,Y, "A5 Gentle start"))
a(("I have to take around 5 of these to get my typical daily dose.", S, A, "Quoted fragment", "Anger", "Simplify my routine", "OBJECTION · Underdosed", "Reading the label", n,n,n,n,Y, "A4 Dose is the product"))
a(("Amazon's sticker was placed directly over the bottle label.", S, A, P, "Anger", "Verify it's real and safe", "OBJECTION · Physical quality", "Unboxing / first inspection", n,n,n,n,Y, "A2 The receipt"))
a(("The product page shows a flip-top lid but the bottle has a regular lid.", S, A, P, "Fear", "Not get ripped off", "OBJECTION · Claim vs reality", "Unboxing / first inspection", n,n,n,n,Y, "A2 The receipt"))
a(("No expiration date on the bottle. I warned others not to buy.", S, A, P, "Anger", "Verify it's real and safe", "OBJECTION · No expiry", "Warning others", n,n,n,n,Y, "A2 The receipt"))

# ---------------- Renue Senolytique ----------------
S = "Renue By Science Senolytique"; A = "Listing summary"
a(("Quercetin at only 77mg when studies typically use 500mg a day.", S, A, P, "Anger", "Be the person who did the research", "OBJECTION · Underdosed", "Reading the label", n,n,n,n,n, "A4 Dose is the product"))
a(("Fisetin at only 35mg when the recommended range is 100 to 500mg.", S, A, P, "Anger", "Be the person who did the research", "OBJECTION · Underdosed", "Reading the label", n,n,n,n,n, "A4 Dose is the product"))
a(("Are these micro-doses actually effective?", S, A, P, "Fear", "Not get ripped off", "OBJECTION · Underdosed", "Reading the label", n,n,n,n,Y, "A4 Dose is the product"))
a(("No perceptible benefit after 10 or more days of use.", S, A, P, "Grief", "Know whether it's working", "OBJECTION · Felt nothing", "Weeks 2–4", Y,Y,n,n,Y, "A1 You won't feel it"))
a(("I was alarmed by passion flower extract in the other ingredients list.", S, A, P, "Fear", "Verify it's real and safe", "OBJECTION · Hidden excipient", "Reading the label", n,n,n,n,Y, "A2 The receipt"))
a(("The bottle appears only one-third full.", S, A, P, "Anger", "Not get ripped off", "OBJECTION · Fill level", "Unboxing / first inspection", n,Y,n,n,Y, "A3 Do the math"))

# ---------------- Renue Restore PM ----------------
S = "Renue By Science Restore PM"; A = "Listing summary"
a(("Zero effects, positive or negative.", S, A, "Quoted fragment", "Grief", "Sleep through the night", "OBJECTION · Felt nothing", "Weeks 2–4", n,n,n,n,Y, "A1 You won't feel it"))
a(("Hit or miss. Great some nights, useless on others. It feels like a placebo.", S, A, "Quoted fragment", "Anger", "Sleep through the night", "OBJECTION · Placebo suspicion", "Weeks 2–4", Y,n,n,n,Y, "A1 You won't feel it"))
a(("I've had chronic insomnia for 18 years and saw no benefit.", S, A, P, "Grief", "Sleep through the night", "OBJECTION · Felt nothing", "Weeks 2–4", Y,Y,n,n,Y, "A1 You won't feel it"))
a(("Buying the two ingredients separately was $20 cheaper with better results.", S, A, P, "Anger", "Justify the ongoing spend", "OBJECTION · Cost per day", "Reorder decision", n,n,n,n,Y, "A3 Do the math"))
a(("The melatonin gives me vivid, dark dreams.", S, A, P, "Fear", "Sleep through the night", "OBJECTION · Side effect", "Week 1", n,Y,n,n,Y, "A5 Gentle start"))
a(("Morning grogginess, particularly for older users.", S, A, P, "Grief", "Sleep through the night", "OBJECTION · Side effect", "Week 1", Y,Y,n,n,Y, "A5 Gentle start"))
a(("Is the third-party lab really independent, or the same lab that made the ingredients?", S, A, P, "Fear", "Not get ripped off", "OBJECTION · Testing credibility", "Pre-purchase research", n,n,n,n,Y, "A2 The receipt"))

# ---------------- Renue Collagen ----------------
S = "Renue By Science Collagen Powder"; A = "Listing summary"
a(("It tasted like chocolate Gatorade, and not in a good way.", S, A, "Quoted fragment", "Anger", "Protect a sensitive body", "OBJECTION · Taste", "First dose", n,Y,n,n,Y, "A5 Gentle start"))
a(("Bitter, sour, acrid and oddly salty.", S, A, P, "Anger", "Protect a sensitive body", "OBJECTION · Taste", "First dose", n,Y,n,n,Y, "A5 Gentle start"))
a(("I poured it down the drain after being unable to finish even one cup.", S, A, P, "Anger", "Protect a sensitive body", "OBJECTION · Taste", "First dose", n,Y,n,Y,Y, "A5 Gentle start"))
a(("Stomach gurgling loud enough that my coworkers noticed.", S, A, P, "Shame", "Protect a sensitive body", "OBJECTION · Gut reaction", "Week 1", Y,Y,n,Y,Y, "A5 Gentle start"))
a(("Splitting it into 3 smaller doses a day helped reduce the symptoms.", S, A, P, "Relief", "Protect a sensitive body", "BENEFIT · Self-invented protocol", "Week 1", n,Y,n,n,Y, "A5 Gentle start"))
a(("It overpowers coffee and tastes unpleasant in water.", S, A, P, "Anger", "Simplify my routine", "OBJECTION · Not versatile", "First dose", n,Y,n,n,Y, "A5 Gentle start"))

# ---------------- Renue LIPO NR ----------------
S = "Renue By Science LIPO NR"; A = "Listing summary"
a(("It works better than Tru Niagen and regular NR supplements.", S, A, P, "Relief", "Get my energy back", "BENEFIT · Beats benchmark", "Comparing brands", n,n,Y,n,Y, "A4 Dose is the product"))

# ================= BUILD =================
wb = openpyxl.Workbook()

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="2E3D31")
HDR_FONT = Font(name=FONT, size=9, bold=True, color="FFFFFF")
BASE = Font(name=FONT, size=10)
QUOTE_FONT = Font(name=FONT, size=10, italic=True)
thin = Side(style="thin", color="D0D4CC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
TOP = Alignment(vertical="top", wrap_text=True)
CTR = Alignment(vertical="top", horizontal="center")

# ---- Sheet 1: the bank ----
ws = wb.active
ws.title = "Quote Bank"
headers = ["ID","Customer's own words","Words","Brand / SKU","Source","Provenance",
           "Emotion","Job to be done","Objection or Benefit","Trigger moment",
           "Time/place","Physical","Former self","Admission","Plain","<12 words",
           "Marker score","Verdict","Angle"]
ws.append(headers)

for i, r in enumerate(R, start=2):
    (q, brand, src, prov, emo, jtbd, ob, trig, t, p, f, e, l, ang) = r
    ws.cell(i, 1, f"Q{i-1:03d}")
    ws.cell(i, 2, q)
    ws.cell(i, 3, f'=IF(LEN(TRIM(B{i}))=0,0,LEN(TRIM(B{i}))-LEN(SUBSTITUTE(TRIM(B{i})," ",""))+1)')
    ws.cell(i, 4, brand); ws.cell(i, 5, src); ws.cell(i, 6, prov)
    ws.cell(i, 7, emo); ws.cell(i, 8, jtbd); ws.cell(i, 9, ob); ws.cell(i, 10, trig)
    ws.cell(i, 11, t); ws.cell(i, 12, p); ws.cell(i, 13, f); ws.cell(i, 14, e); ws.cell(i, 15, l)
    ws.cell(i, 16, f'=IF(C{i}<12,"Y","")')
    ws.cell(i, 17, f'=COUNTIF(K{i}:P{i},"Y")')
    ws.cell(i, 18, f'=IF(Q{i}>=3,"USE","CONTEXT")')
    ws.cell(i, 19, ang)

last = len(R) + 1

widths = {"A":7,"B":58,"C":7,"D":26,"E":25,"F":19,"G":11,"H":27,"I":29,"J":22,
          "K":10,"L":9,"M":11,"N":11,"O":8,"P":9,"Q":8,"R":10,"S":22}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

for row in ws.iter_rows(min_row=1, max_row=last, max_col=19):
    for c in row:
        c.font = BASE; c.alignment = TOP; c.border = BORDER
for c in ws[1]:
    c.font = HDR_FONT; c.fill = HDR_FILL
    c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
ws.row_dimensions[1].height = 34

for i in range(2, last + 1):
    ws.cell(i, 2).font = QUOTE_FONT
    for col in (3, 11, 12, 13, 14, 15, 16, 17, 18):
        ws.cell(i, col).alignment = CTR

# conditional-ish static emphasis
USE_FILL = PatternFill("solid", fgColor="E4EEE6")
CTX_FILL = PatternFill("solid", fgColor="F2F2EE")
score_map = {}
for i, r in enumerate(R, start=2):
    marks = sum(1 for v in r[8:13] if v == "Y")
    wc = len(r[0].split())
    total = marks + (1 if wc < 12 else 0)
    score_map[i] = total
    fill = USE_FILL if total >= 3 else CTX_FILL
    ws.cell(i, 18).fill = fill
    ws.cell(i, 18).font = Font(name=FONT, size=10, bold=True,
                               color="2E6B3E" if total >= 3 else "8A8F86")

# lock the header + first two columns while scrolling
ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:S{last}"

DEAD = PatternFill("solid", fgColor="F7DDD8")
for i, r in enumerate(R, start=2):
    if "do not use" in r[13].lower():
        for col in range(1, 20):
            ws.cell(i, col).fill = DEAD

# ---- Sheet 2: rollups ----
ws2 = wb.create_sheet("Rollups")
ws2["A1"] = "Where the emotional weight sits"
ws2["A1"].font = Font(name=FONT, size=13, bold=True, color="2E3D31")
ws2["A2"] = "All counts are live formulas over the Quote Bank sheet."
ws2["A2"].font = Font(name=FONT, size=9, italic=True, color="777777")

def block(start_row, title, header, values, col_letter):
    ws2.cell(start_row, 1, title).font = Font(name=FONT, size=11, bold=True)
    ws2.cell(start_row + 1, 1, header).font = HDR_FONT
    ws2.cell(start_row + 1, 1).fill = HDR_FILL
    ws2.cell(start_row + 1, 2, "Quotes").font = HDR_FONT
    ws2.cell(start_row + 1, 2).fill = HDR_FILL
    ws2.cell(start_row + 1, 3, "Usable (score 3+)").font = HDR_FONT
    ws2.cell(start_row + 1, 3).fill = HDR_FILL
    for j, v in enumerate(values):
        rr = start_row + 2 + j
        ws2.cell(rr, 1, v).font = BASE
        ws2.cell(rr, 2, f"=COUNTIF('Quote Bank'!${col_letter}$2:${col_letter}${last},A{rr})").font = BASE
        ws2.cell(rr, 3, f"=COUNTIFS('Quote Bank'!${col_letter}$2:${col_letter}${last},A{rr},'Quote Bank'!$R$2:$R${last},\"USE\")").font = BASE
    return start_row + 2 + len(values) + 1

emotions = ["Fear","Shame","Hope","Relief","Pride","Anger","Grief"]
nxt = block(4, "By emotion", "Emotion", emotions, "G")

jtbds = sorted({r[5] for r in R})
nxt = block(nxt + 1, "By job to be done", "Job to be done", jtbds, "H")

trigs = ["Pre-purchase research","Comparing brands","Reading the label","Doing the price math",
         "Unboxing / first inspection","First dose","Week 1","Weeks 2–4","Month 2+",
         "Reorder decision","Abandonment / refund","Warning others"]
nxt = block(nxt + 1, "By trigger moment", "Trigger moment", trigs, "J")

angles = sorted({r[13] for r in R})
block(nxt + 1, "By angle (maps to the teardown)", "Angle", angles, "S")

ws2.column_dimensions["A"].width = 40
ws2.column_dimensions["B"].width = 10
ws2.column_dimensions["C"].width = 18
for row in ws2.iter_rows():
    for c in row:
        if c.alignment.wrap_text is not True:
            c.alignment = Alignment(vertical="center")

# ---- Sheet 3: legend ----
ws3 = wb.create_sheet("How to use this")
lines = [
    ("H1", "Voice-of-Customer Bank — how to use it"),
    ("P", "One quote per row. Tagged so you can filter down to the exact line you need, then write copy that is transcription rather than invention."),
    ("H2", "The six markers"),
    ("P", "A quote earns a marker for each of these it contains. Column Q sums them; column R turns that into a verdict."),
    ("B", "Time/place — names a specific hour, day, age or setting. \"I took two pills around 8am.\""),
    ("B", "Physical — a bodily sensation, not an abstraction. \"Stomach gurgling loud enough that my coworkers noticed.\""),
    ("B", "Former self — measures now against a previous version of themselves. \"At 65 I needed more energy, not less.\""),
    ("B", "Admission — something slightly embarrassing to say out loud. \"I counted them to be sure.\""),
    ("B", "Plain — everyday unbranded speech, no jargon, no ingredient names."),
    ("B", "<12 words — computed live from the word count in column C."),
    ("H2", "The verdict rule"),
    ("P", "3 or more markers = USE. These are the lines that can carry an ad, a subject line, or a headline more or less as written. Under 3 = CONTEXT: true and useful for understanding the buyer, but too abstract to run as copy. Roughly a third of the bank clears the bar, which is normal — you are looking for the few dozen lines that sound like a person, not a category."),
    ("H2", "Provenance — read this before you quote anything publicly"),
    ("B", "Verbatim — transcribed from a full review in the source document. The reviewer name, star rating and date are in column E."),
    ("B", "Quoted fragment — a phrase that appeared in quotation marks inside a listing summary. Real customer wording, but partial and unattributed."),
    ("B", "Paraphrase — not quotable — a summariser's restatement of what reviewers said. Use it to find the angle. Never present it as a customer quote."),
    ("P", "Only Verbatim rows are safe to reproduce as testimonials, and even then they belong to competitors' customers — they are research input for your own copy, not social proof you can publish. Any performance claim you echo needs its own substantiation, and creator content needs its material connection disclosed."),
    ("H2", "Rows shaded red"),
    ("P", "Three rows touch disease treatment or medication replacement (COPD, IBS, blood pressure). They are in the bank because they reveal real buyer intent, and they are flagged because building any claim on them would be a regulatory problem. Serve that intent with gentleness and simplicity messaging only."),
    ("H2", "Source caveat"),
    ("P", "The underlying document mixes about ten full verbatim reviews with AI-generated summaries of competitor listings. Percentages quoted in those summaries are directional, not measured. Treat this bank as a language and hypothesis source, then re-run the same tagging on your own reviews, support tickets and refund reasons — that is where the Prime-specific version lives."),
]
rr = 1
for kind, text in lines:
    c = ws3.cell(rr, 1, text)
    if kind == "H1":
        c.font = Font(name=FONT, size=15, bold=True, color="2E3D31"); rr += 2
    elif kind == "H2":
        c.font = Font(name=FONT, size=11, bold=True, color="2E3D31"); rr += 1
    elif kind == "B":
        c.value = "•  " + text
        c.font = Font(name=FONT, size=10); c.alignment = Alignment(wrap_text=True, vertical="top"); rr += 1
    else:
        c.font = Font(name=FONT, size=10); c.alignment = Alignment(wrap_text=True, vertical="top"); rr += 2
ws3.column_dimensions["A"].width = 118

wb.save("/home/user/Merin/nmn-review-teardown/voice-of-customer-bank.xlsx")
print("rows:", len(R))
print("USE (score>=3):", sum(1 for v in score_map.values() if v >= 3))

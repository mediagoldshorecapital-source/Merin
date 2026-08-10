# -*- coding: utf-8 -*-
import sys
sys.path.insert(0, '/tmp/claude-0/-home-user-Merin/66489eb8-aec6-586f-abae-afced68687e0/scratchpad')
from matrix_data import ROWS, COLUMNS, TRACKER_COLS

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

FONT = "Arial"
INK = "1A1A1A"
HDR_FILL = PatternFill("solid", fgColor="0F3D46")
HDR_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BODY = Font(name=FONT, size=10, color=INK)
BODY_B = Font(name=FONT, size=10, bold=True, color=INK)
TITLE = Font(name=FONT, size=15, bold=True, color="0F3D46")
SUB = Font(name=FONT, size=10, italic=True, color="5A6A72")
FILLIN = PatternFill("solid", fgColor="FFF6CC")
BAND = PatternFill("solid", fgColor="F2F6F7")
ROBERT = PatternFill("solid", fgColor="EAF4F6")
TOP = Alignment(vertical="top", wrap_text=True)
TOP_C = Alignment(vertical="top", horizontal="center", wrap_text=True)
THIN = Side(style="thin", color="D6DEE2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TIER_COLOR = {"Proven": "1B6547", "Adjacent": "8A5A0B", "Untested": "8A2F1E"}

wb = Workbook()

# ---------------------------------------------------------------- Sheet 1
ws = wb.active
ws.title = "How to use"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 104

def line(r, label, text, bold=False, title=False):
    if title:
        ws.cell(r, 2, label).font = TITLE
        return
    if label:
        c = ws.cell(r, 2, label); c.font = BODY_B; c.alignment = TOP
    if text:
        c = ws.cell(r, 3, text); c.font = BODY_B if bold else BODY; c.alignment = TOP

line(2, "NMN Pain-Point → Creative Brief Matrix", None, title=True)
ws.cell(3, 2, "Prime Ingredients · built 10 Aug 2026").font = SUB
ws.merge_cells("B3:C3")

r = 5
line(r, "What this is", "30 customer pain points for NMN Complex, each mapped to a producible ad brief. A designer should be able to build the poster, and an editor cut the reel, from a single row without asking a follow-up question."); r += 2
line(r, "Sheets", "‘Pain-point inventory’ = the 30 statements with surface pain, deep pain and the evidence anchor. ‘Creative matrix’ = the production table."); r += 2
line(r, "Sources", "NMN Deep Customer Intelligence Report (20 May 2026) for the pain points; six-month Meta creative teardown of 1,220 ads / $381K spend for what converts."); r += 2

line(r, "COLUMNS YOU FILL IN", "Owner · Status · Asset link — the three yellow columns at the right of the Creative matrix. Everything else is reference; leave it as written.", bold=True); r += 1
line(r, "", "Example row values →   Owner: “J. Alvarez”   ·   Status: “In edit”   ·   Asset link: “https://drive.google.com/…”"); r += 2

line(r, "Evidence tier", "Proven (8 rows) = the angle has winning ads behind it. Adjacent (10) = a variant of a winner. Untested (12) = new. Spend accordingly — do not launch twelve untested concepts at once."); r += 2
line(r, "Funnel stage", "Cold / Warm / Retention. Respect it. Running a retention concept as cold traffic is how ASTA is currently mispositioned."); r += 2

line(r, "NON-NEGOTIABLE RULES", "", bold=True); r += 1
rules = [
 "1. Hooks qualify, they do not attract. Hook rate, hold rate and CTR all correlate NEGATIVELY with return in this account, in all seven months measured. Every hook must give a non-buyer a reason to scroll past.",
 "2. Committed verbs only — plummeting / collapsing / rebuild. Never ‘naturally declines’, ‘helps support’, ‘may assist’. This is the 0.92-vs-1.41 ROAS difference and it is the most important rule here.",
 "3. Never put ‘75% OFF’ in a headline. It lost on conversions-per-click in five months out of five. Lead with the per-bottle anchor: $29.98 → $16.66.",
 "4. Talent must read 58–68. She compares herself to her peer group; a 45-year-old model breaks the mirror.",
 "5. Reels: 20s, hook inside 3 seconds, captions burned in. She watches in a recliner with the TV on and the sound off.",
 "6. Plain beats polished. In this account the winners look homemade and the losers look like adverts.",
]
for t in rules:
    c = ws.cell(r, 3, t); c.font = BODY; c.alignment = TOP; r += 1
r += 1

line(r, "Judge new ads on", "Conversions per click, NOT hook rate or CTR.   ≥4.0% scale  ·  3.0–4.0% hold and iterate  ·  <2.8% kill.   Give each ad ~40 clicks (≈$60) first. Read it from Meta — the Triple Whale checkout event only fires on ~63% of purchases."); r += 2
line(r, "Naming convention", "NMN_[PainPoint#]_[Angle]_[Format]_[Variant]_[YYYYMM]     e.g.  NMN_16_QualityContrast_DemoCOA_A_202608"); r += 1
line(r, "", "Three separate rows currently exist for what looks like one ad (‘- Video - Ad -2’, ‘- Video - Ad - 2’, ‘Video - Ad - 2’). Adopt this scheme so angle-level reporting works on arrival."); r += 2
line(r, "Compliance flags", "Rows 14 (weight) and 23 (cancellation claim) need review before they run. Row 15 is blocked until Collagen COGS is populated."); r += 2
line(r, "Robert", "Rows 26–30 are the male buyer, reported at 1.81 ROAS — the only segment above the ~1.70 breakeven, and with no creative written to him today. That figure comes from the research, not the warehouse: verify in Ads Manager before scaling.")

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=3):
    for c in row:
        if c.alignment is None or not c.alignment.wrap_text:
            c.alignment = TOP

# ---------------------------------------------------------------- Sheet 2
inv = wb.create_sheet("Pain-point inventory")
inv.sheet_view.showGridLines = False
inv_cols = [("n", "#", 5), ("persona", "Persona", 10), ("cat", "Category", 24),
            ("tier", "Evidence tier", 13), ("pain", "Pain point (her words)", 40),
            ("surface", "Surface pain", 32), ("deep", "Deep pain", 62),
            ("evidence", "Evidence anchor", 58)]
for i, (_, hdr, w) in enumerate(inv_cols, start=1):
    inv.column_dimensions[get_column_letter(i)].width = w
    c = inv.cell(1, i, hdr); c.font = HDR_FONT; c.fill = HDR_FILL
    c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
inv.row_dimensions[1].height = 28

for ri, row in enumerate(ROWS, start=2):
    for ci, (key, _, _) in enumerate(inv_cols, start=1):
        c = inv.cell(ri, ci, row[key])
        c.font = BODY; c.alignment = TOP_C if key in ("n", "persona", "tier") else TOP
        c.border = BORDER
        if row["persona"] == "Robert":
            c.fill = ROBERT
        elif ri % 2 == 0:
            c.fill = BAND
        if key == "tier":
            c.font = Font(name=FONT, size=10, bold=True, color=TIER_COLOR[row["tier"]])
        if key == "pain":
            c.font = BODY_B
    inv.row_dimensions[ri].height = 74
inv.freeze_panes = "E2"
inv.auto_filter.ref = f"A1:{get_column_letter(len(inv_cols))}{len(ROWS)+1}"

# ---------------------------------------------------------------- Sheet 3
mx = wb.create_sheet("Creative matrix")
mx.sheet_view.showGridLines = False
all_cols = COLUMNS + [(f"__t{i}", t, 18) for i, t in enumerate(TRACKER_COLS)]
for i, (_, hdr, w) in enumerate(all_cols, start=1):
    mx.column_dimensions[get_column_letter(i)].width = w
    c = mx.cell(1, i, hdr); c.font = HDR_FONT; c.fill = HDR_FILL
    c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
mx.row_dimensions[1].height = 32

n_ref = len(COLUMNS)
for ri, row in enumerate(ROWS, start=2):
    for ci, (key, _, _) in enumerate(all_cols, start=1):
        val = row.get(key, "")
        c = mx.cell(ri, ci, val)
        c.font = BODY
        c.alignment = TOP_C if key in ("n", "persona", "tier") else TOP
        c.border = BORDER
        if ci > n_ref:
            c.fill = FILLIN
        elif row["persona"] == "Robert":
            c.fill = ROBERT
        elif ri % 2 == 0:
            c.fill = BAND
        if key == "tier":
            c.font = Font(name=FONT, size=10, bold=True, color=TIER_COLOR[row["tier"]])
        if key in ("pain", "hook", "heading"):
            c.font = BODY_B
    mx.row_dimensions[ri].height = 170
mx.freeze_panes = "F2"
mx.auto_filter.ref = f"A1:{get_column_letter(len(all_cols))}{len(ROWS)+1}"

out = "/home/user/Merin/analysis/nmn-pain-point-matrix/nmn-creative-matrix.xlsx"
import os
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("saved", out)
print("rows:", len(ROWS), "| ref cols:", n_ref, "| tracker cols:", len(TRACKER_COLS))
from collections import Counter
print("tiers:", dict(Counter(r["tier"] for r in ROWS)))
print("persona:", dict(Counter(r["persona"] for r in ROWS)))
print("category:", dict(Counter(r["cat"] for r in ROWS)))
missing = [(r["n"], k) for r in ROWS for k, _, _ in COLUMNS if not str(r.get(k, "")).strip()]
print("EMPTY CELLS:", missing if missing else "none")

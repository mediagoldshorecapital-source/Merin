# -*- coding: utf-8 -*-
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from icp_data import ICPS, QUOTES, EMOTIONS, FLAG_LABELS, FINDINGS, HARVEST
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

F = "Arial"
HDR_FILL = PatternFill("solid", fgColor="0F3D46")
HDR = Font(name=F, size=10, bold=True, color="FFFFFF")
BODY = Font(name=F, size=10, color="1A1A1A")
BOLD = Font(name=F, size=10, bold=True, color="1A1A1A")
TITLE = Font(name=F, size=15, bold=True, color="0F3D46")
SUB = Font(name=F, size=10, italic=True, color="5A6A72")
BAND = PatternFill("solid", fgColor="F2F6F7")
YES = PatternFill("solid", fgColor="DCEDE4")
READY = PatternFill("solid", fgColor="FFF6CC")
TOP = Alignment(vertical="top", wrap_text=True)
CTR = Alignment(vertical="top", horizontal="center", wrap_text=True)
THIN = Side(style="thin", color="D6DEE2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
EMO_COLOR = {"Fear": "8A2F1E", "Shame": "6B3F8A", "Hope": "1B6547", "Relief": "1B6547",
             "Pride": "0F6E8A", "Anger": "C03A1E", "Grief": "5A4A3A"}

wb = Workbook()

# ---------------------------------------------------------------- read me
ws = wb.active; ws.title = "Read me"
ws.sheet_view.showGridLines = False
for col, w in (("A", 3), ("B", 26), ("C", 104)):
    ws.column_dimensions[col].width = w
r = 2
ws.cell(r, 2, "Four-ICP Dossier + Voice-of-Customer Quote Bank").font = TITLE; r += 1
ws.cell(r, 2, "Prime Ingredients · built 10 Aug 2026").font = SUB; r += 2

def row(lbl, txt, bold=False):
    global r
    if lbl:
        c = ws.cell(r, 2, lbl); c.font = BOLD; c.alignment = TOP
    c = ws.cell(r, 3, txt); c.font = BOLD if bold else BODY; c.alignment = TOP
    r += 1

row("Sheets", "‘ICP dossier’ = the four personas, all 12 parts. ‘Quote bank’ = verbatim customer language. "
              "‘Findings’ = operational problems customers are reporting publicly. ‘Harvest plan’ = how to grow the bank.")
r += 1
row("CONFIDENCE KEY", "Every claim in the dossier carries a marker. Do not treat them as equivalent.", bold=True)
row("[M] Measured", "From the Triple Whale warehouse or the 468-comment corpus. Trust these.")
row("[R] Research", "From the NMN Deep Customer Intelligence Report, 20 May 2026. Some of it is now out of date — the timing data in particular.")
row("[I] Inferred", "Reasoned from the above. Explicitly not observed. Treat as hypothesis.")
row("[—] Not available", "No data exists. Named rather than guessed.")
r += 1
row("QUOTE BANK RULE", "Every quote is verbatim and traceable to a comment_id in social_media_comments_table. "
                       "Nothing is paraphrased, tidied or invented. Spelling and punctuation are the customer's own.", bold=True)
r += 1
row("The gap, stated", "The corpus is objection-heavy. Fear and anger dominate; HOPE and SHAME return zero rows and "
                       "RELIEF and PRIDE only a handful. That is not an oversight — it is what exists. Those cells are "
                       "empty so you know where to go looking. See the Harvest plan sheet.")
r += 1
row("Copy-ready", "A quote scoring 3 or more of the six quality flags is marked COPY-READY. Those are the lines that "
                  "can go into an ad almost as written.")
r += 1
row("Biggest caveat", "ICP 2 (Linda — Clarity) rests on ManyChat self-selection, which is not in the warehouse. "
                      "It cannot be sized or tied to revenue. It is a creative hypothesis, not a measured segment.")

# ---------------------------------------------------------------- dossier
ds = wb.create_sheet("ICP dossier")
ds.sheet_view.showGridLines = False
part_names = list(ICPS[0]["parts"].keys())
ds.column_dimensions["A"].width = 34
for i in range(len(ICPS)):
    ds.column_dimensions[get_column_letter(2 + i)].width = 62

c = ds.cell(1, 1, "Part"); c.font = HDR; c.fill = HDR_FILL; c.alignment = TOP
for i, icp in enumerate(ICPS):
    c = ds.cell(1, 2 + i, f"{icp['name']}\n{icp['tagline']}")
    c.font = HDR; c.fill = HDR_FILL; c.alignment = TOP
ds.row_dimensions[1].height = 44

meta_rows = [("Size", "size"), ("Evidence strength", "evidence")]
rr = 2
for lbl, key in meta_rows:
    c = ds.cell(rr, 1, lbl); c.font = BOLD; c.alignment = TOP; c.border = BORDER; c.fill = BAND
    for i, icp in enumerate(ICPS):
        c = ds.cell(rr, 2 + i, icp[key]); c.font = BOLD; c.alignment = TOP; c.border = BORDER; c.fill = BAND
    ds.row_dimensions[rr].height = 32
    rr += 1

for p in part_names:
    c = ds.cell(rr, 1, p); c.font = BOLD; c.alignment = TOP; c.border = BORDER
    for i, icp in enumerate(ICPS):
        c = ds.cell(rr, 2 + i, icp["parts"].get(p, "—")); c.font = BODY; c.alignment = TOP; c.border = BORDER
    ds.row_dimensions[rr].height = 170
    rr += 1
ds.freeze_panes = "B2"

# ---------------------------------------------------------------- quote bank
qs = wb.create_sheet("Quote bank")
qs.sheet_view.showGridLines = False
cols = [("Quote (verbatim)", 78), ("Emotion", 11), ("Job-to-be-done", 30), ("Objection / Benefit", 13),
        ("What it objects to / proves", 34), ("Trigger moment", 34), ("ICP", 22),
        ("Copy-ready", 12), ("Flags", 8)] + [(lbl, 15) for _, lbl in FLAG_LABELS] + \
       [("Date", 11), ("comment_id", 34), ("Source", 20)]
for i, (h, w) in enumerate(cols, start=1):
    qs.column_dimensions[get_column_letter(i)].width = w
    c = qs.cell(1, i, h); c.font = HDR; c.fill = HDR_FILL
    c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
qs.row_dimensions[1].height = 40

for ri, q in enumerate(sorted(QUOTES, key=lambda x: (x["kind"] != "Benefit", x["emotion"])), start=2):
    nflags = len(q["flags"])
    ready = "COPY-READY" if nflags >= 3 else ""
    vals = [q["text"], q["emotion"], q["jtbd"], q["kind"], q["obj"], q["trigger"], q["icp"],
            ready, nflags] + ["Yes" if k in q["flags"] else "" for k, _ in FLAG_LABELS] + \
           [q["dt"], q["cid"], "Meta ad comment"]
    for ci, v in enumerate(vals, start=1):
        c = qs.cell(ri, ci, v)
        c.font = BODY; c.alignment = TOP; c.border = BORDER
        if ri % 2 == 0:
            c.fill = BAND
        if ci == 1:
            c.font = BOLD
        if ci == 2:
            c.font = Font(name=F, size=10, bold=True, color=EMO_COLOR.get(q["emotion"], "1A1A1A"))
            c.alignment = CTR
        if ci == 8 and ready:
            c.fill = READY; c.font = Font(name=F, size=10, bold=True, color="8A5A0B"); c.alignment = CTR
        if ci == 9:
            c.alignment = CTR
        if 10 <= ci <= 15 and v == "Yes":
            c.fill = YES; c.alignment = CTR
    qs.row_dimensions[ri].height = 62
qs.freeze_panes = "B2"
qs.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(QUOTES)+1}"

# emotion coverage block
er = len(QUOTES) + 4
qs.cell(er, 1, "EMOTION COVERAGE — where the corpus is empty").font = BOLD
er += 1
counts = {e: sum(1 for q in QUOTES if q["emotion"] == e) for e in EMOTIONS}
for e in EMOTIONS:
    c = qs.cell(er, 1, e); c.font = BODY
    c2 = qs.cell(er, 2, counts[e]); c2.alignment = CTR
    note = "" if counts[e] >= 5 else ("NO DATA — harvest required" if counts[e] == 0 else "THIN — harvest required")
    c3 = qs.cell(er, 3, note); c3.font = Font(name=F, size=10, bold=True, color="8A2F1E" if counts[e] < 5 else "1B6547")
    er += 1

# ---------------------------------------------------------------- findings
fs = wb.create_sheet("Findings")
fs.sheet_view.showGridLines = False
for col, w in (("A", 11), ("B", 46), ("C", 78), ("D", 78), ("E", 52)):
    fs.column_dimensions[col].width = w
for i, h in enumerate(["Severity", "Finding", "What it is", "Customers' own words", "Action"], start=1):
    c = fs.cell(1, i, h); c.font = HDR; c.fill = HDR_FILL; c.alignment = TOP
fs.row_dimensions[1].height = 26
for ri, f in enumerate(FINDINGS, start=2):
    vals = [f["sev"].upper(), f["title"], f["body"], "\n\n".join('"' + x + '"' for x in f["quotes"]), f["action"]]
    for ci, v in enumerate(vals, start=1):
        c = fs.cell(ri, ci, v); c.font = BODY; c.alignment = TOP; c.border = BORDER
        if ri % 2 == 0: c.fill = BAND
        if ci == 1:
            c.font = Font(name=F, size=10, bold=True,
                          color={"critical": "C03A1E", "high": "8A5A0B", "medium": "0F6E8A"}[f["sev"]])
            c.alignment = CTR
        if ci == 2: c.font = BOLD
    fs.row_dimensions[ri].height = 210

# ---------------------------------------------------------------- harvest
hs = wb.create_sheet("Harvest plan")
hs.sheet_view.showGridLines = False
for col, w in (("A", 7), ("B", 34), ("C", 34), ("D", 96), ("E", 30)):
    hs.column_dimensions[col].width = w
for i, h in enumerate(["Rank", "Source", "Estimated volume", "Why it matters", "Effort"], start=1):
    c = hs.cell(1, i, h); c.font = HDR; c.fill = HDR_FILL; c.alignment = TOP
for ri, h in enumerate(HARVEST, start=2):
    for ci, v in enumerate([h["rank"], h["source"], h["est"], h["why"], h["effort"]], start=1):
        c = hs.cell(ri, ci, v); c.font = BODY; c.alignment = CTR if ci == 1 else TOP; c.border = BORDER
        if ri % 2 == 0: c.fill = BAND
        if ci == 2: c.font = BOLD
    hs.row_dimensions[ri].height = 74

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "icp-quote-bank.xlsx")
wb.save(out)
print("saved", out)
print("quotes:", len(QUOTES), "| copy-ready:", sum(1 for q in QUOTES if len(q["flags"]) >= 3))
print("emotion coverage:", counts)
print("ICPs:", len(ICPS), "| parts each:", len(part_names))
